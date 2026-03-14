import argparse
import builtins
import importlib
import pickle as pk
import re
from pathlib import Path
from typing import Dict, List, Optional, Union
import numpy as np
import torch
from PIL import Image

# Optional Transformers imports (only load what you use)
from transformers import CLIPProcessor, CLIPModel
from transformers import ViltProcessor, ViltModel
from transformers import ViTModel, ViTImageProcessor
from transformers import ViTMAEModel
from transformers import AutoTokenizer, AutoModel, AutoImageProcessor


# --------------------------
# Global model for tokenization and dimension reference
model_name = "bert-base-uncased"   # any masked-LM or encoder works; dim stays fixed for a given model
tok = None
mdl = None
# --------------------------
# Helpers
# --------------------------

def word_from_filename(p: Union[str, Path]) -> str:
    name = Path(p).stem
    name = name.replace("_", " ").replace("-", " ").strip()
    return name


def normalize_word_label(word: str) -> str:
    word = str(word).strip().lower()
    word = word.replace("_", " ").replace("-", " ")
    return re.sub(r"\d+$", "", word).strip()

def _to_device(batch: Dict[str, torch.Tensor], device: torch.device) -> Dict[str, torch.Tensor]:
    return {k: v.to(device) if isinstance(v, torch.Tensor) else v for k, v in batch.items()}

def _avg_last_k_layers(hidden_states: List[torch.Tensor], k: int = 4) -> torch.Tensor:
    # hidden_states: list length L, each [B, T, D]; return [B, T, D]
    k = min(k, len(hidden_states))
    return torch.stack(hidden_states[-k:]).mean(0)

def _token_indices_for_span(text: str, word: str, tokenizer) -> List[int]:
    """
    Find token indices that cover the first occurrence of `word` (case-insensitive) in `text`.
    Works with *fast* tokenizers that return offsets.
    """
    text_low = text.lower()
    word_low = word.lower()
    start_char = text_low.find(word_low)
    if start_char < 0:
        return []  # not found; fall back to heuristic if needed

    end_char = start_char + len(word_low)
    tok = tokenizer(text, return_offsets_mapping=True, add_special_tokens=True)
    offsets = tok["offset_mapping"]  # List[(s, e)]
    idxs = []
    for i, (s, e) in enumerate(offsets):
        # Special tokens often have (0,0); skip those
        if e == 0 and s == 0:
            continue
        if s >= start_char and e <= end_char:
            idxs.append(i)
    return idxs


def _pool_clip_vision_layer(hidden_state: torch.Tensor) -> torch.Tensor:
    """Pool a vision transformer hidden state [B, num_patches+1, D] using the CLS token (index 0)."""
    return hidden_state[:, 0, :]  # [B, D]


def _pool_clip_text_layer(hidden_state: torch.Tensor, input_ids: torch.Tensor) -> torch.Tensor:
    """
    Pool a text transformer hidden state [B, seq_len, D] using the EOS token position.
    CLIP uses the EOS token (highest token id in each sequence) as the pooled representation.
    """
    # EOS token is at the position of the largest token id (eos_token_id) in each sequence
    eos_positions = input_ids.argmax(dim=-1)  # [B]
    pooled = hidden_state[torch.arange(hidden_state.size(0)), eos_positions]  # [B, D]
    return pooled


def _pool_feature_tensor(hidden_state: torch.Tensor) -> torch.Tensor:
    """Pool feature maps/sequences into [B, D] vectors for mixed model families."""
    if hidden_state.ndim == 4:
        # CNN-style feature map: [B, C, H, W] -> global average pooling
        return hidden_state.mean(dim=(2, 3))
    if hidden_state.ndim == 3:
        # Transformer-style sequence: [B, T, D] -> CLS token when available
        return hidden_state[:, 0, :]
    if hidden_state.ndim == 2:
        return hidden_state
    raise ValueError(f"Unsupported hidden state shape: {tuple(hidden_state.shape)}")


# --------------------------
# Main embedder
# --------------------------

class MultimodalEmbedder:
    """
    backend='clip'   -> returns layer-wise embeddings from both vision and text encoders
    backend='vit'    -> returns layer-wise embeddings from a pure ImageNet-trained ViT
    backend='mae'    -> returns layer-wise embeddings from a pure MAE vision transformer
    backend='dinov2' -> returns layer-wise embeddings from a pure DINOv2 vision transformer
    backend='simclr' -> returns layer-wise embeddings from a SimCLR vision encoder
    backend='vilt'   -> returns {'word_fused': [D], 'cls_fused': [D]}
    """
    def __init__(
        self,
        backend: str = "clip",
        model_name: Optional[str] = None,
        device: Optional[str] = None,
        fp16: bool = True,
    ):
        self.backend = backend.lower()
        self.device = torch.device(device or ("cuda" if torch.cuda.is_available() else "cpu"))
        self.fp16 = fp16 and self.device.type == "cuda"

        if self.backend == "clip":
            self.model_name = model_name or "openai/clip-vit-base-patch32"
            self.processor = CLIPProcessor.from_pretrained(self.model_name)
            self.model = CLIPModel.from_pretrained(self.model_name).to(self.device).eval()

        elif self.backend == "vit":
            self.model_name = model_name or "google/vit-base-patch32-224-in21k"
            self.processor = ViTImageProcessor.from_pretrained(self.model_name)
            self.model = ViTModel.from_pretrained(self.model_name).to(self.device).eval()

        elif self.backend == "mae":
            self.model_name = model_name or "facebook/vit-mae-base"
            self.processor = AutoImageProcessor.from_pretrained(self.model_name)
            self.model = ViTMAEModel.from_pretrained(self.model_name).to(self.device).eval()

        elif self.backend == "dinov2":
            self.model_name = model_name or "facebook/dinov2-base"
            self.processor = AutoImageProcessor.from_pretrained(self.model_name)
            self.model = AutoModel.from_pretrained(self.model_name).to(self.device).eval()

        elif self.backend == "simclr":
            self.model_name = model_name or "microsoft/resnet-50"
            # Local .pt/.pth SimCLR checkpoint → torchvision ResNet + forward hooks
            self._simclr_local = Path(self.model_name).is_file()
            if self._simclr_local:
                import torchvision.models as tvm
                from torchvision import transforms as T
                encoder = tvm.resnet50(weights=None)
                ckpt = torch.load(self.model_name, map_location="cpu")
                state = ckpt.get("state_dict", ckpt.get("model", ckpt))
                # Strip common SimCLR wrapper prefixes; drop projection head
                state = {
                    k.replace("encoder.", "").replace("backbone.", ""): v
                    for k, v in state.items()
                    if not any(k.startswith(p) for p in ("projector", "projection_head", "fc."))
                }
                encoder.load_state_dict(state, strict=False)
                self.model = encoder.to(self.device).eval()
                self.processor = T.Compose([
                    T.Resize(256),
                    T.CenterCrop(224),
                    T.ToTensor(),
                    T.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225]),
                ])
            else:
                # HuggingFace path — default: microsoft/resnet-50 (public ResNet-50)
                # Pass any other HF ResNet model ID or a local HF directory.
                self.processor = AutoImageProcessor.from_pretrained(self.model_name)
                self.model = AutoModel.from_pretrained(self.model_name).to(self.device).eval()

        elif self.backend == "vilt":
            self.model_name = model_name or "dandelin/vilt-b32-mlm"
            self.processor = ViltProcessor.from_pretrained(self.model_name)
            self.model = ViltModel.from_pretrained(self.model_name).to(self.device).eval()

        elif self.backend == "visualbert":
            raise NotImplementedError(
                "VisualBERT requires object-region features. "
                "If you can provide region features (visual_embeds/boxes), I can wire this up."
            )
        else:
            raise ValueError(f"Unknown backend: {backend}")

    @torch.no_grad()
    def embed_one(self, image_path: Union[str, Path], word: Optional[str] = None,
                  text_template: str = "a photo of {word}"):
        """
        Returns a dict of numpy vectors depending on backend.

        For CLIP, returns:
          - 'vision_layer_00' ... 'vision_layer_12': CLS-pooled embedding at each layer [D_vision]
              Layer 00 = initial patch embedding, layers 01-12 = transformer block outputs
          - 'text_layer_00' ... 'text_layer_12': EOS-pooled embedding at each layer [D_text]
              Layer 00 = initial token embedding, layers 01-12 = transformer block outputs
          - 'vision_projected': final projected + normalized image embedding [D_proj]
          - 'text_projected': final projected + normalized text embedding [D_proj]

        For ViT (pure vision), returns:
          - 'vit_layer_00' ... 'vit_layer_12': CLS-pooled embedding at each layer [D]
              Layer 00 = initial patch embedding, layers 01-12 = transformer block outputs
          - 'vit_pooled': final pooler output [D]

                For MAE (pure vision), returns:
                    - 'mae_layer_00' ... 'mae_layer_12': CLS-pooled embedding at each layer [D]
                            Layer 00 = initial patch embedding, layers 01-12 = transformer block outputs
                    - 'mae_pooled': final pooled embedding [D] (final-layer CLS)

                For DINOv2 (pure vision), returns:
                    - 'dinov2_layer_00' ... 'dinov2_layer_12': CLS-pooled embedding at each layer [D]
                            Layer 00 = initial patch embedding, layers 01-12 = transformer block outputs
                    - 'dinov2_pooled': final pooled embedding [D] (if available)

                For SimCLR (pure vision), returns:
                    - 'simclr_layer_00' ... 'simclr_layer_NN': pooled embedding at each hidden layer [D]
                    - 'simclr_pooled': final pooled embedding [D]
                    Layer source: CNN stage outputs (torchvision) or ResNet stage hidden_states (HuggingFace).
                    Default model: microsoft/resnet-50 (HF). Pass a local .pt/.pth path for a real SimCLR ckpt.
        """
        image_path = Path(image_path)
        word = word or word_from_filename(image_path)
        # Remove numbers and convert letters to lowercase in the word
        word = ''.join([c.lower() for c in word if not c.isdigit()])
        image = Image.open(image_path).convert("RGB")

        if self.backend == "clip":
            text = word
            batch = self.processor(text=[text], images=[image], return_tensors="pt", padding=True)
            batch = _to_device(batch, self.device)

            with torch.autocast(device_type=self.device.type, dtype=torch.float16, enabled=self.fp16):
                out = self.model(**batch, output_hidden_states=True)

            result = {}

            # --- Vision encoder: layer-wise embeddings ---
            # out.vision_model_output.hidden_states is a tuple of (num_layers + 1) tensors
            # each of shape [B, num_patches+1, D_vision]
            vision_hidden = out.vision_model_output.hidden_states
            for i, hs in enumerate(vision_hidden):
                pooled = _pool_clip_vision_layer(hs)  # [1, D_vision]
                result[f"vision_layer_{i:02d}"] = pooled[0].detach().float().cpu().numpy()

            # --- Text encoder: layer-wise embeddings ---
            # out.text_model_output.hidden_states is a tuple of (num_layers + 1) tensors
            # each of shape [B, seq_len, D_text]
            text_hidden = out.text_model_output.hidden_states
            input_ids = batch["input_ids"]
            for i, hs in enumerate(text_hidden):
                pooled = _pool_clip_text_layer(hs, input_ids)  # [1, D_text]
                result[f"text_layer_{i:02d}"] = pooled[0].detach().float().cpu().numpy()

            # --- Final projected embeddings (after projection head + normalization) ---
            img_vec = torch.nn.functional.normalize(out.image_embeds, dim=-1)[0].detach().cpu().numpy()
            txt_vec = torch.nn.functional.normalize(out.text_embeds, dim=-1)[0].detach().cpu().numpy()
            result["vision_projected"] = img_vec
            result["text_projected"] = txt_vec

            return result

        elif self.backend == "vit":
            # Pure vision ViT — no text input needed
            batch = self.processor(images=[image], return_tensors="pt")
            batch = _to_device(batch, self.device)

            with torch.autocast(device_type=self.device.type, dtype=torch.float16, enabled=self.fp16):
                out = self.model(**batch, output_hidden_states=True)

            result = {}

            # Layer-wise embeddings: CLS token (index 0) at each layer
            # out.hidden_states is a tuple of (num_layers + 1) tensors
            # each of shape [B, num_patches+1, D]
            for i, hs in enumerate(out.hidden_states):
                cls_vec = hs[:, 0, :]  # [1, D]
                result[f"vit_layer_{i:02d}"] = cls_vec[0].detach().float().cpu().numpy()

            # Final pooler output (CLS token after the pooling layer)
            if out.pooler_output is not None:
                result["vit_pooled"] = out.pooler_output[0].detach().float().cpu().numpy()

            return result

        elif self.backend == "mae":
            # Pure vision MAE encoder — no text input needed
            batch = self.processor(images=[image], return_tensors="pt")
            batch = _to_device(batch, self.device)

            with torch.autocast(device_type=self.device.type, dtype=torch.float16, enabled=self.fp16):
                out = self.model(**batch, output_hidden_states=True)

            result = {}

            # Layer-wise embeddings: CLS token (index 0) at each layer.
            for i, hs in enumerate(out.hidden_states):
                cls_vec = hs[:, 0, :]  # [1, D]
                result[f"mae_layer_{i:02d}"] = cls_vec[0].detach().float().cpu().numpy()

            # ViT-MAE does not expose a pooler output; use final-layer CLS.
            result["mae_pooled"] = out.last_hidden_state[0, 0, :].detach().float().cpu().numpy()

            return result

        elif self.backend == "dinov2":
            # Pure vision DINOv2 encoder — no text input needed
            batch = self.processor(images=[image], return_tensors="pt")
            batch = _to_device(batch, self.device)

            with torch.autocast(device_type=self.device.type, dtype=torch.float16, enabled=self.fp16):
                out = self.model(**batch, output_hidden_states=True)

            result = {}

            # Layer-wise embeddings: CLS token (index 0) at each layer
            # out.hidden_states is a tuple of (num_layers + 1) tensors
            # each of shape [B, num_patches+1, D]
            for i, hs in enumerate(out.hidden_states):
                cls_vec = hs[:, 0, :]  # [1, D]
                result[f"dinov2_layer_{i:02d}"] = cls_vec[0].detach().float().cpu().numpy()

            # Prefer pooler output when provided by the model, else use final-layer CLS.
            if getattr(out, "pooler_output", None) is not None:
                result["dinov2_pooled"] = out.pooler_output[0].detach().float().cpu().numpy()
            else:
                result["dinov2_pooled"] = out.last_hidden_state[0, 0, :].detach().float().cpu().numpy()

            return result

        elif self.backend == "simclr":
            result = {}
            if self._simclr_local:
                # torchvision ResNet: tap each stage with forward hooks
                x = self.processor(image).unsqueeze(0).to(self.device)
                activations: Dict[str, torch.Tensor] = {}
                hooks = []

                def _make_hook(name: str):
                    def _h(_, __, out_): activations[name] = out_
                    return _h

                children = dict(self.model.named_children())
                stage_names = ["conv1", "layer1", "layer2", "layer3", "layer4"]
                for sn in stage_names:
                    if sn in children:
                        hooks.append(children[sn].register_forward_hook(_make_hook(sn)))

                with torch.autocast(device_type=self.device.type, dtype=torch.float16, enabled=self.fp16):
                    _ = self.model(x)

                for h in hooks:
                    h.remove()

                for i, sn in enumerate(stage_names):
                    if sn in activations:
                        pooled = _pool_feature_tensor(activations[sn])
                        result[f"simclr_layer_{i:02d}"] = pooled[0].detach().float().cpu().numpy()

                # Global average pool before the fc head (2048-d SimCLR representation)
                if "layer4" in activations:
                    feat = self.model.avgpool(activations["layer4"]).flatten(1)
                    result["simclr_pooled"] = feat[0].detach().float().cpu().numpy()
            else:
                # HuggingFace path (e.g. microsoft/resnet-50)
                batch = self.processor(images=[image], return_tensors="pt")
                batch = _to_device(batch, self.device)

                with torch.autocast(device_type=self.device.type, dtype=torch.float16, enabled=self.fp16):
                    out = self.model(**batch, output_hidden_states=True, return_dict=True)

                hidden_states = getattr(out, "hidden_states", None) or []
                for i, hs in enumerate(hidden_states):
                    pooled = _pool_feature_tensor(hs)
                    result[f"simclr_layer_{i:02d}"] = pooled[0].detach().float().cpu().numpy()

                if getattr(out, "pooler_output", None) is not None:
                    pooled_vec = out.pooler_output
                elif getattr(out, "last_hidden_state", None) is not None:
                    pooled_vec = _pool_feature_tensor(out.last_hidden_state)
                elif hidden_states:
                    pooled_vec = _pool_feature_tensor(hidden_states[-1])
                else:
                    raise RuntimeError("SimCLR model did not return hidden_states or a pooled output")

                result["simclr_pooled"] = pooled_vec[0].detach().float().cpu().numpy()

            return result

        elif self.backend == "vilt":
            # Build a short prompt to ensure the target word appears explicitly
            text = text_template.format(word=word)
            batch = self.processor(text=[text], images=[image], return_tensors="pt", padding=True)
            batch = _to_device(batch, self.device)

            with torch.autocast(device_type=self.device.type, dtype=torch.float16, enabled=self.fp16):
                out = self.model(**batch, output_hidden_states=True, return_dict=True)

            # CLS-pooled fused representation
            fused_cls = out.pooler_output[0].detach().float().cpu().numpy()

            # Word-level fused embedding: average last 4 layers over the subword span
            last4 = _avg_last_k_layers(out.hidden_states, k=4)  # [1, T, D]
            # Re-tokenize to get offsets (same tokenizer as processor)
            idxs = _token_indices_for_span(text, word, self.processor.tokenizer)
            if not idxs:
                ids = self.processor.tokenizer(text, add_special_tokens=True)["input_ids"]
                idxs = [i for i, tid in enumerate(ids) if tid not in (101, 102)][:1]

            word_vec = last4[0, idxs].mean(0).detach().float().cpu().numpy()
            return {"word_fused": word_vec, "cls_fused": fused_cls}

    @torch.no_grad()
    def embed_folder(
        self,
        folder: Union[str, Path],
        patterns: tuple = ("*.jpg", "*.jpeg", "*.png", "*.webp", "*.bmp"),
        text_template: str = "a photo of {word}",
    ) -> Dict[str, np.ndarray]:
        """
        Iterates a folder, returns dict with:
          - 'paths': np.ndarray of file paths [N]
          - 'words': np.ndarray of word labels [N]
          - For CLIP: 'vision_layer_00' ... 'vision_layer_12' each [N, D_vision]
                      'text_layer_00' ... 'text_layer_12' each [N, D_text]
                      'vision_projected', 'text_projected' each [N, D_proj]
          - For ViT:  'vit_layer_00' ... 'vit_layer_12' each [N, D]
                      'vit_pooled' [N, D]
                    - For MAE:  'mae_layer_00' ... 'mae_layer_12' each [N, D]
                                            'mae_pooled' [N, D]
                    - For DINOv2: 'dinov2_layer_00' ... 'dinov2_layer_12' each [N, D]
                                                 'dinov2_pooled' [N, D]
                    - For SimCLR: 'simclr_layer_00' ... 'simclr_layer_NN' each [N, D]
                                                 'simclr_pooled' [N, D]
                                 HF default: microsoft/resnet-50. Pass a local .pt/.pth for a SimCLR ckpt.
          - For ViLT: 'word_fused' [N, D], 'cls_fused' [N, D]
        """
        folder = Path(folder)
        files = []
        for pat in patterns:
            files.extend(folder.rglob(pat))
        files = sorted(files)

        paths: List[str] = []
        words: List[str] = []
        per_key: Dict[str, List[np.ndarray]] = {}

        for p in files:
            w = word_from_filename(p)
            vecs = self.embed_one(p, word=w, text_template=text_template)

            paths.append(str(p))
            words.append(w.lower())
            for k, v in vecs.items():
                per_key.setdefault(k, []).append(v)

        # Stack
        for k in per_key:
            per_key[k] = np.stack(per_key[k], axis=0)  # [N, D]

        return {"paths": np.array(paths), "words": np.array(words), **per_key}


class PWESuitePanphonEmbedder:
    def __init__(
        self,
        checkpoint_path: Optional[Union[str, Path]] = None,
        device: Optional[str] = None,
        dimension: int = 300,
        feature_mode: str = "panphon",
        ipa_vocab_path: Optional[Union[str, Path]] = None,
    ):
        self.device = torch.device(device or ("cuda" if torch.cuda.is_available() else "cpu"))
        self.dimension = dimension
        self.feature_mode = feature_mode
        if self.feature_mode not in {"panphon", "token_ipa"}:
            raise ValueError(f"Unsupported PWESuite feature mode: {self.feature_mode}")

        default_checkpoint = "rnn_metric_learning_panphon_all.pt" if self.feature_mode == "panphon" else "rnn_metric_learning_token_ipa_all.pt"
        self.checkpoint_path = Path(
            checkpoint_path
            or Path(__file__).parent / f"supportive_repos/pwesuite/computed/models/{default_checkpoint}"
        )

        if not self.checkpoint_path.exists():
            raise FileNotFoundError(f"PWESuite checkpoint not found: {self.checkpoint_path}")

        self.ipa_backend = ""
        self.feature_backend = ""
        self._initialize_pronunciation_backend()

        import sys as _sys
        _pwesuite_root = str(Path(__file__).parent / "supportive_repos/pwesuite")
        if _pwesuite_root not in _sys.path:
            _sys.path.insert(0, _pwesuite_root)
        from models.metric_learning.model import RNNMetricLearner

        state_dict = torch.load(self.checkpoint_path, map_location="cpu")

        # Recover checkpoint input dimensionality from the first LSTM weight matrix.
        first_weight_key = next(k for k in state_dict.keys() if "weight_ih_l0" in k)
        self.input_size = int(state_dict[first_weight_key].shape[1])

        if self.feature_mode == "panphon" and self.input_size != 24:
            raise ValueError(
                f"Panphon mode expects input_size=24, got input_size={self.input_size} from {self.checkpoint_path}"
            )

        if self.feature_mode == "token_ipa":
            self._initialize_ipa_vocab(ipa_vocab_path=ipa_vocab_path)

        self.model = RNNMetricLearner(dimension=dimension, feature_size=self.input_size)
        self.model.load_state_dict(state_dict)
        self.model.to(self.device).eval()

    @staticmethod
    def _utf8_open(*args, **kwargs):
        if "b" not in kwargs.get("mode", ""):
            kwargs.setdefault("encoding", "utf-8")
        return builtins.open(*args, **kwargs)

    def _import_panphon_with_utf8(self):
        original_open = builtins.open

        def _open_utf8(file, mode="r", *args, **kwargs):
            if "b" not in mode:
                kwargs.setdefault("encoding", "utf-8")
            return original_open(file, mode, *args, **kwargs)

        builtins.open = _open_utf8
        try:
            panphon = importlib.import_module("panphon")
            return panphon, panphon.FeatureTable()
        finally:
            builtins.open = original_open

    def _initialize_pronunciation_backend(self) -> None:
        try:
            import eng_to_ipa
            self.epi = eng_to_ipa
            self.ipa_backend = "eng_to_ipa"

            if self.feature_mode == "panphon":
                _, feature_table = self._import_panphon_with_utf8()
                self.ft = feature_table
                self.feature_backend = "panphon"
            else:
                self.ft = None
                self.feature_backend = "ipa_chars"
            return
        except Exception as exc:
            raise RuntimeError(
                "Could not initialize a pronunciation backend. "
                "Install panphon+epitran, or use eng_to_ipa+panphon on Windows."
            ) from exc

    def _initialize_ipa_vocab(self, ipa_vocab_path: Optional[Union[str, Path]] = None) -> None:
        vocab_path = Path(
            ipa_vocab_path
            or Path(__file__).parent / "supportive_repos/pwesuite/data/vocab/ipa_multi.txt"
        )
        if not vocab_path.exists():
            raise FileNotFoundError(
                "IPA feature mode requires a vocabulary file. "
                f"Expected at: {vocab_path}"
            )

        symbols = [line.strip() for line in vocab_path.read_text(encoding="utf-8").splitlines() if line.strip()]
        if not symbols:
            raise ValueError(f"IPA vocabulary file is empty: {vocab_path}")

        # Some released checkpoints can have larger input dims than the checked-in vocab file.
        # Pad deterministically so one-hot shape always matches checkpoint input size.
        if len(symbols) < self.input_size:
            symbols = symbols + [f"__extra_{i}__" for i in range(self.input_size - len(symbols))]
        elif len(symbols) > self.input_size:
            symbols = symbols[:self.input_size]

        self.ipa_vocab_path = vocab_path
        self.ipa_vocab = {symbol: idx for idx, symbol in enumerate(symbols)}
        self.ipa_unk_idx = self.ipa_vocab.get("😕", 0)

    def _ipa_segments(self, ipa_text: str) -> List[str]:
        if self.ft is not None and hasattr(self.ft, "ipa_segs"):
            return list(self.ft.ipa_segs(ipa_text))
        return list(ipa_text)

    def word_to_ipa(self, word: str) -> str:
        normalized_word = normalize_word_label(word)
        if self.ipa_backend == "epitran":
            transliterated = self.epi.transliterate(normalized_word)
        else:
            transliterated = self.epi.convert(normalized_word).replace(" ", "")

        if self.ft is not None and hasattr(self.ft, "ipa_segs"):
            ipa = "".join(self.ft.ipa_segs(transliterated))
        else:
            ipa = transliterated
        ipa = re.sub(r"[,ˌˈ]", "", ipa)
        if not ipa:
            raise ValueError(f"Could not derive IPA for word: {word}")
        return ipa

    def word_to_panphon_features(self, word: str) -> np.ndarray:
        ipa = self.word_to_ipa(word)
        features = np.asarray(self.ft.word_to_vector_list(ipa, numeric=True), dtype=np.float32)
        if features.size == 0:
            raise ValueError(f"Panphon feature extraction returned no features for word: {word}")
        # panphon returns ternary {-1, 0, 1}; the model was trained with
        # panphon2 which collapses {-1, 0} → 0 via max(val, 0).
        np.clip(features, 0, 1, out=features)
        return features

    def word_to_token_ipa_features(self, word: str) -> np.ndarray:
        ipa = self.word_to_ipa(word)
        segments = self._ipa_segments(ipa)
        if not segments:
            raise ValueError(f"No IPA segments found for word: {word}")

        indices = [self.ipa_vocab.get(seg, self.ipa_unk_idx) for seg in segments]
        features = np.zeros((len(indices), self.input_size), dtype=np.float32)
        features[np.arange(len(indices)), indices] = 1.0
        return features

    @torch.inference_mode()
    def embed_words(self, words: List[str], batch_size: int = 1) -> Dict[str, np.ndarray]:
        raw_words = [str(word).strip() for word in words]
        normalized_words = [normalize_word_label(word) for word in raw_words]
        ipa_words = [self.word_to_ipa(word) for word in raw_words]

        if batch_size < 1:
            raise ValueError("batch_size must be >= 1")

        embeddings_out: List[np.ndarray] = []
        for i in range(0, len(raw_words), batch_size):
            batch_words = raw_words[i:i + batch_size]
            if self.feature_mode == "panphon":
                batch_features = [self.word_to_panphon_features(word) for word in batch_words]
            else:
                batch_features = [self.word_to_token_ipa_features(word) for word in batch_words]

            embeddings_out.append(self.model.forward(batch_features).detach().float().cpu().numpy())

        embeddings = np.concatenate(embeddings_out, axis=0)

        return {
            "words": np.array([word.lower() for word in raw_words]),
            "normalized_words": np.array(normalized_words),
            "ipa": np.array(ipa_words),
            "phoneme_embedding": embeddings,
        }


def load_words_from_excel(excel_path: Union[str, Path]) -> List[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook.active
    words = []
    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        value = row[0]
        if value is None:
            continue
        value = str(value).strip()
        if value:
            words.append(value)

    if not words:
        raise ValueError(f"No words found in first spreadsheet column: {excel_path}")
    return words


def save_pickle(payload: Dict[str, np.ndarray], output_path: Union[str, Path]) -> None:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("wb") as handle:
        pk.dump(payload, handle)


def extract_picture_naming_vision_embeddings(
    data_folder: Union[str, Path],
    embedding_folder: Union[str, Path],
    targets: List[str],
) -> None:
    embedding_folder = Path(embedding_folder)
    vision_jobs = {
        "clip": ("openai/clip-vit-base-patch32", "clip_layerwise_embeddings.pk"),
        "vit": ("google/vit-base-patch32-224-in21k", "vit_imagenet_layerwise_embeddings.pk"),
        "mae": ("facebook/vit-mae-base", "mae_layerwise_embeddings.pk"),
        "simclr": ("microsoft/resnet-50", "simclr_layerwise_embeddings.pk"),
        "dinov2": ("facebook/dinov2-base", "dinov2_layerwise_embeddings.pk"),
    }

    for backend in targets:
        model_name, output_name = vision_jobs[backend]
        print("=" * 50)
        print(f"Extracting {backend.upper()} embeddings...")
        print("=" * 50)
        embedder = MultimodalEmbedder(backend=backend, model_name=model_name)
        result = embedder.embed_folder(data_folder)

        print(f"Keys in {backend.upper()} result:")
        for key, value in result.items():
            if isinstance(value, np.ndarray):
                print(f"  {key}: shape={value.shape}, dtype={value.dtype}")

        save_pickle(result, embedding_folder / output_name)
        print(f"Saved to {output_name}\n")

        del embedder
        if torch.cuda.is_available():
            torch.cuda.empty_cache()


def extract_picture_naming_phoneme_embeddings(
    excel_path: Union[str, Path],
    embedding_folder: Union[str, Path],
    checkpoint_path: Optional[Union[str, Path]] = None,
    feature_mode: str = "panphon",
    ipa_vocab_path: Optional[Union[str, Path]] = None,
    batch_size: int = 32,
) -> Path:
    words = load_words_from_excel(excel_path)
    embedder = PWESuitePanphonEmbedder(
        checkpoint_path=checkpoint_path,
        feature_mode=feature_mode,
        ipa_vocab_path=ipa_vocab_path,
    )
    result = embedder.embed_words(words, batch_size=batch_size)
    result["source_spreadsheet"] = str(Path(excel_path))
    result["checkpoint_path"] = str(embedder.checkpoint_path)
    result["feature_mode"] = feature_mode
    result["ipa_backend"] = embedder.ipa_backend
    result["feature_backend"] = embedder.feature_backend
    if feature_mode == "token_ipa":
        result["ipa_vocab_path"] = str(embedder.ipa_vocab_path)

    output_name = "pwesuite_panphon_embeddings.pk" if feature_mode == "panphon" else "pwesuite_token_ipa_embeddings.pk"
    output_path = Path(embedding_folder) / output_name
    save_pickle(result, output_path)

    print("=" * 50)
    print(f"Extracting PWESuite {feature_mode} phoneme embeddings...")
    print("=" * 50)
    for key, value in result.items():
        if isinstance(value, np.ndarray):
            print(f"  {key}: shape={value.shape}, dtype={value.dtype}")
    print(f"Saved to {output_path.name}")

    return output_path


# --------------------------
# Example: get word embedding in context using BERT
# --------------------------
def word_embedding_in_context(sentence, target_word):
    global tok, mdl
    if tok is None or mdl is None:
        tok = AutoTokenizer.from_pretrained(model_name)
        mdl = AutoModel.from_pretrained(model_name)

    enc = tok(sentence, return_tensors="pt")
    with torch.no_grad():
        out = mdl(**enc, output_hidden_states=True)
    H = out.last_hidden_state[0]

    tokens = tok.convert_ids_to_tokens(enc["input_ids"][0])
    positions = [i for i,t in enumerate(tokens) if t.replace("##","") == target_word]
    if not positions:
        positions = [i for i,t in enumerate(tokens) if target_word in t.replace("##","")]
    v = torch.mean(H[positions], dim=0) if positions else torch.mean(H, dim=0)
    return v.numpy()


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--targets",
        nargs="+",
        choices=["clip", "vit", "mae", "simclr", "dinov2", "phoneme", "all-vision", "all"],
        default=["clip", "vit", "mae", "simclr", "dinov2"],
    )
    parser.add_argument(
        "--data-folder",
        default=Path(__file__).parent / "data/pictureNaming extended all",
        type=Path,
    )
    parser.add_argument(
        "--wordset-excel",
        default=Path(__file__).parent / "data/wordset picture naming expanded.xlsx",
        type=Path,
    )
    parser.add_argument(
        "--embedding-folder",
        default=Path(__file__).parent / "embeddings/pictureNaming extended all",
        type=Path,
    )
    parser.add_argument(
        "--pwesuite-checkpoint",
        default=None,
        type=Path,
    )
    parser.add_argument(
        "--pwesuite-feature-mode",
        choices=["panphon", "token_ipa"],
        default="panphon",
    )
    parser.add_argument(
        "--pwesuite-ipa-vocab",
        default=Path(__file__).parent / "supportive_repos/pwesuite/data/vocab/ipa_multi.txt",
        type=Path,
    )
    parser.add_argument(
        "--pwesuite-batch-size",
        type=int,
        default=1,
    )
    args = parser.parse_args()

    targets = args.targets
    if "all" in targets:
        targets = ["clip", "vit", "mae", "simclr", "dinov2", "phoneme"]
    elif "all-vision" in targets:
        targets = ["clip", "vit", "mae", "simclr", "dinov2"]

    vision_targets = [target for target in targets if target in {"clip", "vit", "mae", "simclr", "dinov2"}]
    if vision_targets:
        extract_picture_naming_vision_embeddings(
            data_folder=args.data_folder,
            embedding_folder=args.embedding_folder,
            targets=vision_targets,
        )

    if "phoneme" in targets:
        extract_picture_naming_phoneme_embeddings(
            excel_path=args.wordset_excel,
            embedding_folder=args.embedding_folder,
            checkpoint_path=args.pwesuite_checkpoint,
            feature_mode=args.pwesuite_feature_mode,
            ipa_vocab_path=args.pwesuite_ipa_vocab,
            batch_size=args.pwesuite_batch_size,
        )